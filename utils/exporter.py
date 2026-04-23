"""
exporter.py — Export des résultats en Excel + JSON
Génère le fichier de sortie final avec toutes les fiches produits.
"""

import json
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Couleurs
COLOR_WINNER    = "C6EFCE"   # vert clair
COLOR_POTENTIAL = "FFEB9C"   # jaune clair
COLOR_REJECT    = "FFC7CE"   # rouge clair
COLOR_HEADER    = "1F4E79"   # bleu foncé
COLOR_SUBHEADER = "2E75B6"   # bleu moyen


class ProductExporter:

    def export_all(self, products: list[dict], config: dict) -> str:
        """Exporte en Excel + JSON et retourne le chemin du fichier Excel."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        niche     = config.get("niche", "recherche").replace(" ", "_")
        base_path = Path(f"results_{niche}_{timestamp}")

        # JSON complet (données brutes)
        json_path = base_path.with_suffix(".json")
        self._export_json(products, config, json_path)
        logger.info(f"JSON exporté : {json_path}")

        # Excel formaté
        xlsx_path = base_path.with_suffix(".xlsx")
        self._export_excel(products, config, xlsx_path)
        logger.info(f"Excel exporté : {xlsx_path}")

        return str(xlsx_path)

    # ──────────────────────────────────────────────
    # EXPORT JSON
    # ──────────────────────────────────────────────
    def _export_json(self, products: list, config: dict, path: Path):
        data = {
            "session"    : config,
            "generated"  : datetime.now().isoformat(),
            "total"      : len(products),
            "products"   : products,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    # ──────────────────────────────────────────────
    # EXPORT EXCEL
    # ──────────────────────────────────────────────
    def _export_excel(self, products: list, config: dict, path: Path):
        wb = Workbook()

        # Onglet 1 : Tableau récapitulatif
        self._sheet_summary(wb, products, config)

        # Onglet 2 : Fiches produits détaillées
        self._sheet_detail(wb, products)

        # Onglet 3 : Données fournisseurs
        self._sheet_suppliers(wb, products)

        # Supprimer la feuille par défaut
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(path)

    # ──────────────────────────────────────────────
    # ONGLET 1 — RÉSUMÉ
    # ──────────────────────────────────────────────
    def _sheet_summary(self, wb: Workbook, products: list, config: dict):
        ws = wb.create_sheet("📊 Résumé", 0)

        # En-tête session
        ws.merge_cells("A1:L1")
        ws["A1"] = f"🏆 PRODUCT FINDER — Résultats : {config.get('niche', '')} | Marchés : {', '.join(config.get('markets', []))} | {datetime.now().strftime('%d/%m/%Y')}"
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # En-têtes colonnes
        headers = [
            "#", "Produit", "Statut", "Score /190", "%",
            "Agent1\nTendances", "Agent2\nSocial",
            "Agent3\nMarge", "Agent4\nConcurrence",
            "Prix vente $", "Marge nette %",
            "Action recommandée"
        ]
        widths = [4, 30, 14, 10, 6, 10, 10, 10, 12, 11, 12, 30]

        for col, (header, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 35

        # Lignes produits
        for i, product in enumerate(products, 1):
            row = i + 2
            fs  = product.get("final_score", {})
            a3  = product.get("agent3", {})
            margin = a3.get("margin", {})

            status = fs.get("status", "")
            color  = (
                COLOR_WINNER    if "GAGNANT"   in status else
                COLOR_POTENTIAL if "POTENTIEL" in status else
                COLOR_REJECT
            )

            values = [
                i,
                product.get("product", ""),
                status,
                fs.get("total", 0),
                f"{fs.get('percentage', 0)}%",
                fs.get("breakdown", {}).get("agent1_trends", {}).get("score", 0),
                fs.get("breakdown", {}).get("agent2_social", {}).get("score", 0),
                fs.get("breakdown", {}).get("agent3_suppliers", {}).get("score", 0),
                fs.get("breakdown", {}).get("agent4_competition", {}).get("score", 0),
                margin.get("sale_price"),
                f"{margin.get('net_margin_pct', '')}%" if margin.get("net_margin_pct") else "",
                fs.get("action", ""),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill      = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = self._thin_border()
                if col == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            ws.row_dimensions[row].height = 20

        # Freeze header
        ws.freeze_panes = "A3"

    # ──────────────────────────────────────────────
    # ONGLET 2 — FICHES DÉTAILLÉES
    # ──────────────────────────────────────────────
    def _sheet_detail(self, wb: Workbook, products: list):
        ws = wb.create_sheet("📋 Fiches Produits")

        row = 1
        for product in products:
            fs  = product.get("final_score", {})
            a1  = product.get("agent1", {})
            a2  = product.get("agent2", {})
            a3  = product.get("agent3", {})
            a4  = product.get("agent4", {})

            status = fs.get("status", "")
            color  = (
                COLOR_WINNER    if "GAGNANT"   in status else
                COLOR_POTENTIAL if "POTENTIEL" in status else
                COLOR_REJECT
            )

            # Titre produit
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = f"{'─'*3} #{products.index(product)+1} — {product.get('product', '').upper()}"
            ws[f"A{row}"].font      = Font(bold=True, size=12, color="FFFFFF")
            ws[f"A{row}"].fill      = PatternFill("solid", fgColor=COLOR_HEADER)
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 25
            row += 1

            # Données
            margin = a3.get("margin", {})
            rows_data = [
                ("SCORE FINAL", f"{fs.get('total', 0)}/190 ({fs.get('percentage', 0)}%) — {status}"),
                ("Action", fs.get("action", "")),
                ("", ""),
                ("💰 FINANCES", ""),
                ("Prix vente recommandé", f"${margin.get('sale_price', 'N/A')}"),
                ("Coût fournisseur", f"${margin.get('supplier_cost', 'N/A')}"),
                ("Marge nette estimée", f"{margin.get('net_margin_pct', 'N/A')}% (${margin.get('net_margin_usd', 'N/A')})"),
                ("Livraison estimée", f"${margin.get('shipping', 'N/A')}"),
                ("Budget pub estimé", f"${margin.get('ads_cost', 'N/A')}"),
                ("", ""),
                ("📦 FOURNISSEURS", ""),
            ]

            # Fournisseurs
            suppliers = a3.get("suppliers", {})
            for platform, data in suppliers.items():
                if data:
                    rows_data.append((
                        f"  {platform.title()}",
                        f"${data.get('price', 'N/A')} — Note: {data.get('rating', 'N/A')} — {data.get('url', '')}"
                    ))

            rows_data += [
                ("", ""),
                ("📈 TENDANCES", ""),
            ]

            # Google Trends
            trends = a1.get("google_trends", {})
            for market, data in trends.items():
                rows_data.append((f"  Google Trends {market}", data))

            rows_data += [
                ("", ""),
                ("📣 SOCIAL & PUBS", ""),
                ("Facebook Ads — annonceurs", a2.get("facebook", {}).get("unique_advertisers", 0)),
                ("TikTok — vidéos virales", a2.get("tiktok", {}).get("viral_videos_count", 0)),
                ("Formats créatifs", ", ".join(a2.get("tiktok", {}).get("creative_formats", []))),
                ("", ""),
                ("Liens TikTok", ""),
            ]

            for link in a2.get("tiktok", {}).get("top_links", [])[:3]:
                rows_data.append(("  →", link))

            rows_data += [
                ("Liens Facebook Ads", ""),
            ]
            for link in a2.get("facebook", {}).get("sample_links", [])[:3]:
                rows_data.append(("  →", link))

            rows_data += [
                ("", ""),
                ("🔍 CONCURRENCE", ""),
                ("Fourchette prix marché", str(a4.get("amazon_competition", {}).get("price_range", ""))),
                ("Marques présentes", ", ".join(a4.get("amazon_competition", {}).get("brands", [])[:4])),
                ("Faiblesses concurrents", " | ".join(a4.get("pain_points", [])[:2])),
                ("Angle différenciation", " | ".join(a4.get("differentiation", [])[:2])),
                ("", ""),
                ("", "─" * 80),
                ("", ""),
            ]

            for label, value in rows_data:
                ws.cell(row=row, column=1, value=label).font = Font(bold=bool(label and label.isupper() or label.startswith("💰") or label.startswith("📦") or label.startswith("📈") or label.startswith("📣") or label.startswith("🔍")))
                ws.cell(row=row, column=2, value=str(value) if value else "")
                if label and label.isupper():
                    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
                row += 1

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 80

    # ──────────────────────────────────────────────
    # ONGLET 3 — FOURNISSEURS
    # ──────────────────────────────────────────────
    def _sheet_suppliers(self, wb: Workbook, products: list):
        ws = wb.create_sheet("📦 Fournisseurs")

        headers = ["Produit", "Plateforme", "Prix $", "Note", "Commandes", "Délai (j)", "URL"]
        widths  = [30, 18, 10, 8, 12, 10, 60]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font  = Font(bold=True, color="FFFFFF")
            cell.fill  = PatternFill("solid", fgColor=COLOR_SUBHEADER)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = w

        row = 2
        for product in products:
            suppliers = product.get("agent3", {}).get("suppliers", {})
            for platform, data in suppliers.items():
                if data:
                    ws.cell(row=row, column=1, value=product.get("product", ""))
                    ws.cell(row=row, column=2, value=data.get("name", platform))
                    ws.cell(row=row, column=3, value=data.get("price"))
                    ws.cell(row=row, column=4, value=data.get("rating"))
                    ws.cell(row=row, column=5, value=data.get("orders"))
                    ws.cell(row=row, column=6, value=data.get("shipping"))
                    ws.cell(row=row, column=7, value=data.get("url", ""))
                    row += 1

        ws.freeze_panes = "A2"

    def _thin_border(self):
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)
